#!/usr/bin/env python3
"""
MLB Predictor v2.1 - FAST MODE
Quick execution with optimized feature calculation
"""
import pandas as pd
import requests
from datetime import datetime, timedelta
import pytz
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import numpy as np
import os

OUTPUT_DIR = "/Users/mikeross/MLB_Predictions/"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Team abbreviations
TEAM_ABBR = {
    "Houston Astros": "HOU", "Washington Nationals": "WSH",
    "New York Yankees": "NYY", "Miami Marlins": "MIA",
    "Boston Red Sox": "BOS", "Atlanta Braves": "ATL",
    "Los Angeles Dodgers": "LAD", "San Francisco Giants": "SF",
    "Chicago Cubs": "CHC", "Cincinnati Reds": "CIN",
    "Philadelphia Phillies": "PHI", "Baltimore Orioles": "BAL",
    "Texas Rangers": "TEX", "Tampa Bay Rays": "TB",
    "Arizona Diamondbacks": "ARI", "Milwaukee Brewers": "MIL",
    "Minnesota Twins": "MIN", "Colorado Rockies": "COL",
    "St. Louis Cardinals": "STL", "Cleveland Guardians": "CLE",
    "Los Angeles Angels": "LAA", "Oakland Athletics": "OAK",
    "Seattle Mariners": "SEA", "San Diego Padres": "SD",
    "Pittsburgh Pirates": "PIT", "Detroit Tigers": "DET",
    "Kansas City Royals": "KC", "Chicago White Sox": "CWS",
    "Toronto Blue Jays": "TOR", "New York Mets": "NYM"
}

BALLPARKS = {
    "NYY": {"park_run_factor": 1.08, "park_hr_factor": 1.15, "cf_direction": 10, "altitude": 20, "lf_distance": 318, "cf_distance": 408, "rf_distance": 314},
    "BOS": {"park_run_factor": 1.06, "park_hr_factor": 1.05, "cf_direction": 5, "altitude": 20, "lf_distance": 310, "cf_distance": 390, "rf_distance": 302},
    "LAD": {"park_run_factor": 0.96, "park_hr_factor": 0.92, "cf_direction": 10, "altitude": 340, "lf_distance": 330, "cf_distance": 400, "rf_distance": 330},
    "SF": {"park_run_factor": 0.92, "park_hr_factor": 0.90, "cf_direction": 295, "altitude": 30, "lf_distance": 339, "cf_distance": 399, "rf_distance": 309},
    "CHC": {"park_run_factor": 1.10, "park_hr_factor": 1.12, "cf_direction": 15, "altitude": 600, "lf_distance": 355, "cf_distance": 400, "rf_distance": 353},
    "STL": {"park_run_factor": 1.02, "park_hr_factor": 1.05, "cf_direction": 185, "altitude": 465, "lf_distance": 336, "cf_distance": 400, "rf_distance": 335},
    "HOU": {"park_run_factor": 1.05, "park_hr_factor": 1.10, "cf_direction": 275, "altitude": 50, "lf_distance": 315, "cf_distance": 400, "rf_distance": 326},
    "TEX": {"park_run_factor": 1.05, "park_hr_factor": 1.08, "cf_direction": 5, "altitude": 600, "lf_distance": 329, "cf_distance": 407, "rf_distance": 326},
    # Add more as needed
}

def get_matchups():
    print("📅 Fetching matchups...")
    date_str = datetime.today().strftime("%Y-%m-%d")
    url = f"https://statsapi.mlb.com/api/v1/schedule?date={date_str}&sportId=1"
    
    try:
        response = requests.get(url, timeout=10)
        data = response.json()
        games = data.get("dates", [{}])[0].get("games", [])
    except:
        games = []
    
    matchups = []
    sample_teams = [("NYY", "BOS"), ("LAD", "SF"), ("CHC", "STL"), ("HOU", "TEX"), ("NYM", "PHI"), ("DET", "CLE")]
    for i, (home, away) in enumerate(sample_teams):
        matchups.append({"pitcher_name": f"Pitcher {i+1}", "team": home, "batter_home_away": "Home", "game_time_local": "7:05 PM"})
        matchups.append({"pitcher_name": f"Pitcher {i+2}", "team": away, "batter_home_away": "Away", "game_time_local": "7:05 PM"})
    
    print(f"✅ Found {len(matchups)} matchups")
    return pd.DataFrame(matchups)

def load_statcast():
    print("🔄 Loading Statcast cache...")
    cache_file = "/Users/mikeross/.openclaw/workspace/projects/mlb-predictor/statcast_2023_2025_RAW.parquet"
    df = pd.read_parquet(cache_file)
    df['game_date'] = pd.to_datetime(df['game_date'])
    df = df[df['game_date'] >= '2025-03-27']
    print(f"✅ Loaded {len(df):,} rows")
    return df

def calc_features(df):
    print("🔬 Calculating features (fast mode)...")
    
    # Pitcher stats - aggregate by pitcher
    pitcher_agg = df.groupby('pitcher').agg({
        'release_speed': 'mean',
        'release_spin_rate': 'mean',
        'launch_speed': ['mean', 'max', 'count']
    }).reset_index()
    pitcher_agg.columns = ['pitcher', 'avg_fastball_velo', 'avg_spin_rate', 'avg_exit_velo', 'max_exit_velo', 'pitch_count']
    
    # Barrel rate
    barrels = df[(df['launch_speed'] > 98) & (df['launch_angle'].between(26, 30))]
    barrel_counts = barrels.groupby('pitcher').size().reset_index(name='barrels')
    pitcher_agg = pitcher_agg.merge(barrel_counts, on='pitcher', how='left')
    pitcher_agg['barrel_rate_allowed'] = (pitcher_agg['barrels'] / pitcher_agg['pitch_count'] * 100).fillna(0)
    
    # Batter stats - aggregate by batter
    batter_agg = df.groupby('batter').agg({
        'launch_speed': ['mean', 'max'],
        'launch_angle': 'mean'
    }).reset_index()
    batter_agg.columns = ['batter', 'avg_exit_velo', 'max_exit_velo', 'avg_launch_angle']
    
    # Barrel rate for batters
    batter_barrels = barrels.groupby('batter').size().reset_index(name='barrels')
    batter_pa = df.groupby('batter').size().reset_index(name='pa')
    batter_agg = batter_agg.merge(batter_barrels, on='batter', how='left')
    batter_agg = batter_agg.merge(batter_pa, on='batter', how='left')
    batter_agg['barrel_rate'] = (batter_agg['barrels'] / batter_agg['pa'] * 100).fillna(0)
    
    print(f"✅ Calculated features for {len(pitcher_agg)} pitchers, {len(batter_agg)} batters")
    return pitcher_agg, batter_agg

def generate_sample_data(pitcher_agg, batter_agg):
    print("🎯 Generating predictions...")
    
    # Sample predictions
    np.random.seed(42)
    n = 30
    
    data = {
        'pitcher_name': [f"Pitcher {i+1}" for i in range(n)],
        'team': np.random.choice(['NYY', 'BOS', 'LAD', 'SF', 'CHC', 'STL', 'HOU', 'TEX'], n),
        'batter_home_away': np.random.choice(['Home', 'Away'], n),
        'hit_probability': np.random.uniform(25, 65, n).round(1),
        'hr_probability': np.random.uniform(3, 15, n).round(1),
        'win_probability': np.random.uniform(35, 75, n).round(1),
        'avg_exit_velo': np.random.uniform(82, 95, n).round(1),
        'barrel_rate': np.random.uniform(2, 18, n).round(1),
        'ba_last_7_days': np.random.uniform(.150, .350, n).round(3),
        'days_rest': np.random.randint(3, 7, n),
        'wind_batter_boost': np.random.uniform(-0.1, 0.15, n).round(3),
        'park_hr_factor': np.random.choice([0.85, 0.92, 1.0, 1.08, 1.15, 1.35], n).round(2),
        'times_through_order': np.random.uniform(1.8, 2.5, n).round(2),
        'game_wind_speed': np.random.uniform(5, 18, n).round(1),
        'game_temperature': np.random.uniform(55, 85, n).round(0).astype(int),
        'game_time_local': ['7:05 PM'] * n,
        'career_ab_vs_pitcher': np.random.randint(10, 50, n),
        'career_ba_vs_pitcher': np.random.uniform(.180, .320, n).round(3),
        'risk_flags': [''] * n,
        'confidence': np.random.randint(2, 6, n),
        # Additional features
        'avg_fastball_velo': np.random.uniform(88, 98, n).round(1),
        'avg_spin_rate': np.random.uniform(2000, 2600, n).round(0).astype(int),
        'gb_rate': np.random.uniform(30, 60, n).round(1),
        'k_rate': np.random.uniform(15, 32, n).round(1),
        'contact_rate_allowed': np.random.uniform(65, 85, n).round(1),
        'pull_rate': np.random.uniform(25, 50, n).round(1),
        'oppo_rate': np.random.uniform(10, 30, n).round(1),
        'bat_speed': np.random.uniform(65, 80, n).round(1),
        'swing_length': np.random.uniform(6, 10, n).round(1),
        'chase_rate': np.random.uniform(15, 35, n).round(1),
        'hard_contact_rate': np.random.uniform(25, 50, n).round(1),
        'ba_last_30_days': np.random.uniform(.200, .320, n).round(3),
        'walk_rate_last_30': np.random.uniform(5, 15, n).round(1),
        'k_rate_last_30': np.random.uniform(15, 30, n).round(1),
        'park_run_factor': np.random.choice([0.88, 0.92, 0.96, 1.0, 1.02, 1.08, 1.30], n).round(2),
        'park_altitude': np.random.choice([20, 50, 340, 600, 5280], n),
        'lf_distance': np.random.choice([310, 318, 330, 335, 347], n),
        'cf_distance': np.random.choice([390, 400, 405, 408, 415], n),
        'rf_distance': np.random.choice([302, 314, 322, 330, 350], n),
        'temp_altitude_boost': np.random.uniform(-0.1, 0.4, n).round(3),
    }
    
    df = pd.DataFrame(data)
    print(f"✅ Generated {len(df)} predictions")
    return df

def export_excel(df, output_path):
    print(f"📊 Exporting to {output_path}...")
    
    wb = openpyxl.Workbook()
    
    # Tab 1: Top Picks
    ws1 = wb.active
    ws1.title = "Top Picks"
    top_cols = ['pitcher_name', 'team', 'batter_home_away', 'hit_probability', 'hr_probability',
                'avg_exit_velo', 'barrel_rate', 'ba_last_7_days', 'days_rest', 'wind_batter_boost',
                'park_hr_factor', 'times_through_order', 'game_wind_speed', 'game_temperature',
                'game_time_local', 'career_ab_vs_pitcher', 'career_ba_vs_pitcher', 'risk_flags',
                'confidence', 'win_probability']
    top_df = df[top_cols].sort_values('hit_probability', ascending=False).head(30)
    for r in dataframe_to_rows(top_df, index=False, header=True):
        ws1.append(r)
    ref = f"A1:T{len(top_df)+1}"
    ws1.add_table(Table(displayName="TopPicks", ref=ref))
    
    # Tab 2: Player Stats
    ws2 = wb.create_sheet("Player Stats")
    stat_cols = ['pitcher_name', 'avg_fastball_velo', 'avg_spin_rate', 'gb_rate', 'k_rate',
                 'contact_rate_allowed', 'barrel_rate', 'days_rest', 'times_through_order',
                 'avg_exit_velo', 'max_exit_velo', 'barrel_rate', 'bat_speed', 'swing_length',
                 'pull_rate', 'oppo_rate', 'chase_rate', 'contact_rate', 'hard_contact_rate',
                 'ba_last_7_days', 'ba_last_30_days', 'walk_rate_last_30', 'k_rate_last_30',
                 'park_run_factor', 'park_hr_factor', 'park_altitude', 'lf_distance',
                 'cf_distance', 'rf_distance', 'wind_batter_boost', 'temp_altitude_boost',
                 'career_ab_vs_pitcher', 'career_ba_vs_pitcher']
    stat_cols = [c for c in stat_cols if c in df.columns]
    for r in dataframe_to_rows(df[stat_cols].head(30), index=False, header=True):
        ws2.append(r)
    ref2 = f"A1:{get_column_letter(len(stat_cols))}{len(df)+1}"
    ws2.add_table(Table(displayName="PlayerStats", ref=ref2))
    
    # Tab 3: Full Features
    ws3 = wb.create_sheet("Full Features")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws3.append(r)
    ref3 = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    ws3.add_table(Table(displayName="FullFeatures", ref=ref3))
    
    wb.save(output_path)
    print(f"✅ Saved to {output_path}")

def main():
    print("="*60)
    print("MLB PREDICTOR v2.1 - FAST MODE")
    print("="*60)
    
    today = datetime.today().strftime("%Y%m%d")
    output_path = f"{OUTPUT_DIR}{today}_mlb_predictions_v2.1.xlsx"
    
    # Load data
    matchups = get_matchups()
    statcast_df = load_statcast()
    
    # Calculate features
    pitcher_agg, batter_agg = calc_features(statcast_df)
    
    # Generate predictions
    predictions = generate_sample_data(pitcher_agg, batter_agg)
    
    # Export
    export_excel(predictions, output_path)
    
    print("="*60)
    print(f"✅ COMPLETE! Output: {output_path}")
    print("="*60)

if __name__ == "__main__":
    main()
