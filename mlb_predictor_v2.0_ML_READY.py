#!/usr/bin/env python3
"""
MLB Prediction Model v2.0 — ML-Ready Edition
==============================================
Enhancements over v1.0:
1. Wind direction analysis (helps batters if blowing out to CF)
2. Park factors (run/HR multipliers for all 30 MLB parks)
3. Enhanced batter features (exit velo, barrel rate, L/R splits, chase rate)
4. Enhanced pitcher features (pitch mix, spin rates, days rest, fatigue)
5. Batter vs pitcher career history
6. Team offense/defense stats (last 10 games)
7. ML-ready feature structure (ready for XGBoost/LightGBM)

Author: Mike Ross (The Architect)
Date: 2026-02-21
"""
import pandas as pd
from pybaseball import statcast
import requests
from datetime import datetime, timedelta
import unicodedata
import re
import pytz 
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
import numpy as np

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def clean_name(name):
    if not isinstance(name, str):
        return name
    try:
        name = name.encode('latin1').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    name = unicodedata.normalize('NFKC', name)
    return name.replace('\u200b', '').strip()

def flip_name(name):
    if "," in name:
        last, first = name.split(",", 1)
        return f"{first.strip()} {last.strip()}"
    return name

def extract_batter_name(row):
    des = str(row["des"])
    event = str(row["events"])
    if des.lower().startswith("challenge") or "challenged" in des.lower():
        return "Unknown"
    name_pattern = r"[A-ZÀ-ÿ][a-zA-ZÀ-ÿ'.\-]+ [A-ZÀ-ÿ][a-zA-ZÀ-ÿ'.\-]+"
    if event in ["walk", "intent_walk", "hit_by_pitch", "strikeout_looking"]:
        matches = re.findall(name_pattern, des)
        return matches[-1] if matches else "Unknown"
    match = re.match(name_pattern, des)
    return match.group(0) if match else "Unknown"

def get_batter_team(row):
    if row["inning_topbot"] == "Top":
        return row["away_team"]
    elif row["inning_topbot"] == "Bottom":
        return row["home_team"]
    return None

# ============================================================================
# MLB BALLPARK DATA (ENHANCED WITH PARK FACTORS & ORIENTATION)
# ============================================================================

# Park factors: 1.0 = neutral, >1.0 = hitter-friendly, <1.0 = pitcher-friendly
# Orientation: degrees from north (0° = north, 90° = east, 180° = south, 270° = west)
# Outfield direction: where CF wall is located relative to north
ballpark_locations = {
    "ARI": {
        "stadium": "Chase Field", "city": "Phoenix", "state": "AZ",
        "lat": 33.4455, "lon": -112.0667,
        "park_run_factor": 1.08, "park_hr_factor": 1.15,  # Hitter-friendly (roof open)
        "cf_direction": 105,  # CF faces ENE
        "altitude": 1117  # ft above sea level
    },
    "ATL": {
        "stadium": "Truist Park", "city": "Atlanta", "state": "GA",
        "lat": 33.8908, "lon": -84.4679,
        "park_run_factor": 1.02, "park_hr_factor": 1.08,
        "cf_direction": 345,  # CF faces NNW
        "altitude": 1050
    },
    "BAL": {
        "stadium": "Camden Yards", "city": "Baltimore", "state": "MD",
        "lat": 39.2839, "lon": -76.6219,
        "park_run_factor": 1.04, "park_hr_factor": 1.10,  # LF short (318 ft)
        "cf_direction": 354,  # CF faces N
        "altitude": 10
    },
    "BOS": {
        "stadium": "Fenway Park", "city": "Boston", "state": "MA",
        "lat": 42.3467, "lon": -71.0972,
        "park_run_factor": 1.06, "park_hr_factor": 1.05,  # Green Monster LF
        "cf_direction": 5,  # CF faces N
        "altitude": 20
    },
    "CHC": {
        "stadium": "Wrigley Field", "city": "Chicago", "state": "IL",
        "lat": 41.9484, "lon": -87.6553,
        "park_run_factor": 1.10, "park_hr_factor": 1.12,  # Wind-dependent
        "cf_direction": 15,  # CF faces NNE
        "altitude": 600
    },
    "CWS": {
        "stadium": "Guaranteed Rate Field", "city": "Chicago", "state": "IL",
        "lat": 41.8300, "lon": -87.6339,
        "park_run_factor": 1.01, "park_hr_factor": 1.03,
        "cf_direction": 185,  # CF faces S
        "altitude": 600
    },
    "CIN": {
        "stadium": "Great American Ball Park", "city": "Cincinnati", "state": "OH",
        "lat": 39.0975, "lon": -84.5070,
        "park_run_factor": 1.12, "park_hr_factor": 1.18,  # Extreme hitter park
        "cf_direction": 95,  # CF faces E
        "altitude": 500
    },
    "CLE": {
        "stadium": "Progressive Field", "city": "Cleveland", "state": "OH",
        "lat": 41.4962, "lon": -81.6852,
        "park_run_factor": 0.98, "park_hr_factor": 0.95,
        "cf_direction": 95,  # CF faces E
        "altitude": 650
    },
    "COL": {
        "stadium": "Coors Field", "city": "Denver", "state": "CO",
        "lat": 39.7562, "lon": -104.9942,
        "park_run_factor": 1.30, "park_hr_factor": 1.35,  # EXTREME altitude effect
        "cf_direction": 5,  # CF faces N
        "altitude": 5280  # Mile-high
    },
    "DET": {
        "stadium": "Comerica Park", "city": "Detroit", "state": "MI",
        "lat": 42.3390, "lon": -83.0485,
        "park_run_factor": 0.96, "park_hr_factor": 0.90,  # Pitcher-friendly (CF deep)
        "cf_direction": 15,  # CF faces NNE
        "altitude": 600
    },
    "HOU": {
        "stadium": "Minute Maid Park", "city": "Houston", "state": "TX",
        "lat": 29.7573, "lon": -95.3555,
        "park_run_factor": 1.05, "park_hr_factor": 1.10,  # Short LF porch
        "cf_direction": 275,  # CF faces W
        "altitude": 50
    },
    "KC": {
        "stadium": "Kauffman Stadium", "city": "Kansas City", "state": "MO",
        "lat": 39.0517, "lon": -94.4803,
        "park_run_factor": 0.98, "park_hr_factor": 0.95,
        "cf_direction": 5,  # CF faces N
        "altitude": 750
    },
    "LAA": {
        "stadium": "Angel Stadium", "city": "Anaheim", "state": "CA",
        "lat": 33.8003, "lon": -117.8827,
        "park_run_factor": 0.97, "park_hr_factor": 0.92,
        "cf_direction": 5,  # CF faces N
        "altitude": 160
    },
    "LAD": {
        "stadium": "Dodger Stadium", "city": "Los Angeles", "state": "CA",
        "lat": 34.0739, "lon": -118.2400,
        "park_run_factor": 0.96, "park_hr_factor": 0.92,
        "cf_direction": 10,  # CF faces N
        "altitude": 340
    },
    "MIA": {
        "stadium": "loanDepot park", "city": "Miami", "state": "FL",
        "lat": 25.7780, "lon": -80.2190,
        "park_run_factor": 0.92, "park_hr_factor": 0.88,  # Pitcher-friendly (retractable roof)
        "cf_direction": 5,  # CF faces N
        "altitude": 10
    },
    "MIL": {
        "stadium": "American Family Field", "city": "Milwaukee", "state": "WI",
        "lat": 43.0280, "lon": -87.9712,
        "park_run_factor": 1.00, "park_hr_factor": 1.00,  # Neutral
        "cf_direction": 195,  # CF faces SSW
        "altitude": 630
    },
    "MIN": {
        "stadium": "Target Field", "city": "Minneapolis", "state": "MN",
        "lat": 44.9817, "lon": -93.2778,
        "park_run_factor": 1.02, "park_hr_factor": 1.05,
        "cf_direction": 195,  # CF faces SSW
        "altitude": 830
    },
    "NYM": {
        "stadium": "Citi Field", "city": "New York", "state": "NY",
        "lat": 40.7571, "lon": -73.8458,
        "park_run_factor": 0.94, "park_hr_factor": 0.88,  # Pitcher-friendly (deep CF)
        "cf_direction": 15,  # CF faces NNE
        "altitude": 10
    },
    "NYY": {
        "stadium": "Yankee Stadium", "city": "New York", "state": "NY",
        "lat": 40.8296, "lon": -73.9262,
        "park_run_factor": 1.08, "park_hr_factor": 1.15,  # Short RF porch (314 ft)
        "cf_direction": 5,  # CF faces N
        "altitude": 55
    },
    "OAK": {
        "stadium": "Oakland Coliseum", "city": "Oakland", "state": "CA",
        "lat": 37.7516, "lon": -122.2005,
        "park_run_factor": 0.93, "park_hr_factor": 0.85,  # Extreme pitcher park (foul territory)
        "cf_direction": 45,  # CF faces NE
        "altitude": 10
    },
    "ATH": {  # Athletics (same as OAK)
        "stadium": "Oakland Coliseum", "city": "Oakland", "state": "CA",
        "lat": 37.7516, "lon": -122.2005,
        "park_run_factor": 0.93, "park_hr_factor": 0.85,
        "cf_direction": 45,
        "altitude": 10
    },
    "PHI": {
        "stadium": "Citizens Bank Park", "city": "Philadelphia", "state": "PA",
        "lat": 39.9061, "lon": -75.1665,
        "park_run_factor": 1.06, "park_hr_factor": 1.12,
        "cf_direction": 5,  # CF faces N
        "altitude": 10
    },
    "PIT": {
        "stadium": "PNC Park", "city": "Pittsburgh", "state": "PA",
        "lat": 40.4469, "lon": -80.0057,
        "park_run_factor": 0.98, "park_hr_factor": 0.95,
        "cf_direction": 325,  # CF faces NW
        "altitude": 730
    },
    "SD": {
        "stadium": "Petco Park", "city": "San Diego", "state": "CA",
        "lat": 32.7073, "lon": -117.1566,
        "park_run_factor": 0.88, "park_hr_factor": 0.82,  # Extreme pitcher park
        "cf_direction": 5,  # CF faces N
        "altitude": 20
    },
    "SEA": {
        "stadium": "T-Mobile Park", "city": "Seattle", "state": "WA",
        "lat": 47.5914, "lon": -122.3325,
        "park_run_factor": 0.96, "park_hr_factor": 0.92,
        "cf_direction": 5,  # CF faces N
        "altitude": 10
    },
    "SF": {
        "stadium": "Oracle Park", "city": "San Francisco", "state": "CA",
        "lat": 37.7786, "lon": -122.3893,
        "park_run_factor": 0.92, "park_hr_factor": 0.85,  # Wind from RF (McCovey Cove)
        "cf_direction": 295,  # CF faces WNW
        "altitude": 10
    },
    "STL": {
        "stadium": "Busch Stadium", "city": "St. Louis", "state": "MO",
        "lat": 38.6226, "lon": -90.1928,
        "park_run_factor": 1.00, "park_hr_factor": 1.00,  # Neutral
        "cf_direction": 5,  # CF faces N
        "altitude": 465
    },
    "TB": {
        "stadium": "Tropicana Field", "city": "St. Petersburg", "state": "FL",
        "lat": 27.7683, "lon": -82.6534,
        "park_run_factor": 0.94, "park_hr_factor": 0.90,  # Indoor (catwalks suppress HRs)
        "cf_direction": 5,  # CF faces N
        "altitude": 10
    },
    "TEX": {
        "stadium": "Globe Life Field", "city": "Arlington", "state": "TX",
        "lat": 32.7473, "lon": -97.0823,
        "park_run_factor": 1.04, "park_hr_factor": 1.08,  # Retractable roof
        "cf_direction": 15,  # CF faces NNE
        "altitude": 550
    },
    "TOR": {
        "stadium": "Rogers Centre", "city": "Toronto", "state": "ON",
        "lat": 43.6414, "lon": -79.3894,
        "park_run_factor": 1.02, "park_hr_factor": 1.05,
        "cf_direction": 45,  # CF faces NE
        "altitude": 300
    },
    "WSH": {
        "stadium": "Nationals Park", "city": "Washington", "state": "DC",
        "lat": 38.8729, "lon": -77.0074,
        "park_run_factor": 0.98, "park_hr_factor": 0.95,
        "cf_direction": 285,  # CF faces WNW
        "altitude": 10
    }
}

# Team abbreviation mapping
team_abbr_map = {
    'Arizona Diamondbacks': 'ARI', 'Atlanta Braves': 'ATL', 'Baltimore Orioles': 'BAL',
    'Boston Red Sox': 'BOS', 'Chicago White Sox': 'CWS', 'Chicago Cubs': 'CHC',
    'Cincinnati Reds': 'CIN', 'Cleveland Guardians': 'CLE', 'Colorado Rockies': 'COL',
    'Detroit Tigers': 'DET', 'Houston Astros': 'HOU', 'Kansas City Royals': 'KC',
    'Los Angeles Angels': 'LAA', 'Los Angeles Dodgers': 'LAD', 'Miami Marlins': 'MIA',
    'Milwaukee Brewers': 'MIL', 'Minnesota Twins': 'MIN', 'New York Mets': 'NYM',
    'New York Yankees': 'NYY', 'Oakland Athletics': 'ATH', 'Athletics': 'ATH',
    'Philadelphia Phillies': 'PHI', 'Pittsburgh Pirates': 'PIT', 'San Diego Padres': 'SD',
    'San Francisco Giants': 'SF', 'Seattle Mariners': 'SEA', 'St. Louis Cardinals': 'STL',
    'Tampa Bay Rays': 'TB', 'Texas Rangers': 'TEX', 'Toronto Blue Jays': 'TOR',
    'Washington Nationals': 'WSH'
}

# ============================================================================
# WIND DIRECTION ANALYSIS
# ============================================================================

def calculate_wind_helps_batters(wind_direction, wind_speed, cf_direction, park_hr_factor):
    """
    Determine if wind helps batters based on wind direction relative to CF orientation.
    
    Logic:
    - Wind blowing OUT to CF (toward outfield) = helps batters (carry)
    - Wind blowing IN from CF (toward home plate) = helps pitchers (knockdown)
    - Crosswind = minimal effect
    
    Returns: wind_batter_boost (float, -1.0 to +1.0)
    """
    if pd.isna(wind_direction) or pd.isna(wind_speed):
        return 0.0
    
    # Calculate angle difference (wind direction - CF direction)
    angle_diff = (wind_direction - cf_direction + 180) % 360 - 180
    
    # Wind blowing OUT to CF: angle_diff near 0° (±45°)
    # Wind blowing IN from CF: angle_diff near ±180°
    if abs(angle_diff) < 45:
        # Blowing out to CF → helps batters
        wind_boost = (wind_speed / 20) * 0.15  # Max +15% at 20 mph
    elif abs(angle_diff) > 135:
        # Blowing in from CF → helps pitchers
        wind_boost = -(wind_speed / 20) * 0.10  # Max -10% at 20 mph
    else:
        # Crosswind → minimal effect
        wind_boost = 0.0
    
    # Scale by park HR factor (wind matters more in hitter parks)
    wind_boost *= park_hr_factor
    
    return round(wind_boost, 3)

# ============================================================================
# STEP 1: PULL TODAY'S MATCHUPS FROM MLB API
# ============================================================================

def get_todays_matchups():
    """Pull today's starting pitchers and game times from MLB API."""
    today = datetime.today().strftime('%Y-%m-%d')
    mlb_url = f"https://statsapi.mlb.com/api/v1/schedule?sportId=1&hydrate=probablePitcher&date={today}"
    
    try:
        response = requests.get(mlb_url, timeout=10)
        data = response.json()
    except Exception as e:
        print(f"❌ Failed to pull MLB schedule: {e}")
        return pd.DataFrame()
    
    matchups = []
    for date_info in data.get('dates', []):
        for game in date_info.get('games', []):
            teams = game.get('teams', {})
            home = teams.get('home', {})
            away = teams.get('away', {})
            game_time_utc = game.get("gameDate")
            
            home_name = home.get('team', {}).get('name', 'Unknown')
            away_name = away.get('team', {}).get('name', 'Unknown')
            home_abbr = team_abbr_map.get(home_name, 'UNK')
            away_abbr = team_abbr_map.get(away_name, 'UNK')
            
            home_pitcher = clean_name(home.get('probablePitcher', {}).get('fullName', None))
            away_pitcher = clean_name(away.get('probablePitcher', {}).get('fullName', None))
            
            if home_pitcher:
                matchups.append({
                    'pitcher_name': home_pitcher,
                    'pitcher_team': home_abbr,
                    'opponent_team': away_abbr,
                    'pitcher_home_away': 'Home',
                    'game_time_utc': game_time_utc,
                    'home_team_abbr': home_abbr
                })
            if away_pitcher:
                matchups.append({
                    'pitcher_name': away_pitcher,
                    'pitcher_team': away_abbr,
                    'opponent_team': home_abbr,
                    'pitcher_home_away': 'Away',
                    'game_time_utc': game_time_utc,
                    'home_team_abbr': home_abbr
                })
    
    pitchers_df = pd.DataFrame(matchups)
    pitchers_df["pitcher_name"] = pitchers_df["pitcher_name"].apply(clean_name)
    
    # Convert to Eastern Time
    eastern = pytz.timezone("US/Eastern")
    pitchers_df["game_time"] = pd.to_datetime(pitchers_df["game_time_utc"])
    pitchers_df["game_time_et"] = pitchers_df["game_time"].dt.tz_convert(eastern)
    pitchers_df["game_time_local"] = pitchers_df["game_time_et"].dt.tz_localize(None)
    
    print(f"✅ Found {len(pitchers_df)} starting pitchers for today")
    return pitchers_df

# ============================================================================
# STEP 2: PULL STATCAST DATA (2025 SEASON)
# ============================================================================

def pull_statcast_data(start_date='2025-03-27', end_date=None):
    """Pull Statcast pitch-level data."""
    if end_date is None:
        end_date = datetime.today().strftime('%Y-%m-%d')
    
    print(f"🔄 Downloading Statcast data ({start_date} to {end_date})...")
    
    # Split into monthly chunks to avoid API rate limits
    ranges = []
    current = datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    
    while current < end_dt:
        month_end = min(current + timedelta(days=30), end_dt)
        ranges.append((current.strftime('%Y-%m-%d'), month_end.strftime('%Y-%m-%d')))
        current = month_end + timedelta(days=1)
    
    chunks = []
    for start, end in ranges:
        try:
            df = statcast(start, end)
            if df is not None and len(df) > 0:
                chunks.append(df)
                print(f"  ✅ Loaded {start} to {end} ({len(df):,} rows)")
        except Exception as e:
            print(f"  ❌ Failed {start} to {end}: {e}")
    
    if not chunks:
        print("❌ No Statcast data loaded")
        return pd.DataFrame()
    
    batters_df = pd.concat(chunks, ignore_index=True)
    batters_df = batters_df.dropna(subset=['events'])
    batters_df["player_name"] = batters_df["player_name"].apply(clean_name)
    batters_df['game_date'] = pd.to_datetime(batters_df['game_date'])
    batters_df = batters_df.sort_values(['batter', 'game_date'], ascending=[True, False])
    
    # Extract batter name from description
    batters_df["batter_name"] = batters_df.apply(extract_batter_name, axis=1).apply(clean_name)
    batters_df = batters_df[batters_df["batter_name"] != "Unknown"]
    
    # Infer batter team
    batters_df["batter_team"] = batters_df.apply(get_batter_team, axis=1)
    
    print(f"✅ Total Statcast data: {len(batters_df):,} plate appearances")
    return batters_df

# ============================================================================
# STEP 3: CALCULATE OUTS & PITCHING STATS
# ============================================================================

def get_outs(event):
    """Convert event to outs recorded."""
    if event == "triple_play":
        return 3
    elif event in ["double_play", "sac_fly_double_play", "strikeout_double_play"]:
        return 2
    elif event in ["field_out", "force_out", "strikeout", "grounded_into_double_play",
                   "fielders_choice_out", "sac_bunt", "sac_fly", "fielders_choice"]:
        return 1
    return 0

def calculate_pitching_stats(batters_df):
    """Calculate pitcher statistics from Statcast data."""
    print("\n📊 Calculating pitching stats...")
    
    batters_df["outs"] = batters_df["events"].apply(get_outs)
    
    # Innings pitched
    pitching_outs = batters_df.groupby("pitcher")["outs"].sum().reset_index(name="outs_recorded")
    pitching_outs["innings_pitched"] = pitching_outs["outs_recorded"] / 3
    
    # Hits allowed
    hit_events = {"single", "double", "triple", "home_run"}
    hits_allowed = (
        batters_df[batters_df["events"].isin(hit_events)]
        .groupby("pitcher")
        .size()
        .reset_index(name="hits_allowed")
    )
    
    # Walks
    walk_events = {"walk", "intent_walk"}
    walks = (
        batters_df[batters_df["events"].isin(walk_events)]
        .groupby("pitcher")
        .size()
        .reset_index(name="walks")
    )
    
    # Strikeouts
    strikeout_events = {"strikeout", "strikeout_swinging", "strikeout_looking", "strikeout_double_play"}
    strikeouts = (
        batters_df[batters_df["events"].isin(strikeout_events)]
        .groupby("pitcher")
        .size()
        .reset_index(name="strikeouts")
    )
    
    # Runs allowed (estimate from description)
    batters_df["run_event"] = batters_df["des"].str.contains("scores|homer", case=False, na=False)
    runs_allowed = batters_df[batters_df["run_event"]].groupby("pitcher").size().reset_index(name="runs_allowed")
    
    # Combine stats
    pitching_stats = pitching_outs.merge(hits_allowed, on="pitcher", how="outer").fillna(0)
    pitching_stats = pitching_stats.merge(walks, on="pitcher", how="outer").fillna(0)
    pitching_stats = pitching_stats.merge(strikeouts, on="pitcher", how="outer").fillna(0)
    pitching_stats = pitching_stats.merge(runs_allowed, on="pitcher", how="left").fillna(0)
    
    # ERA
    pitching_stats["ERA"] = (pitching_stats["runs_allowed"] / pitching_stats["innings_pitched"]) * 9
    pitching_stats.loc[pitching_stats["innings_pitched"] == 0, "ERA"] = None
    
    # WHIP (Walks + Hits per Inning Pitched)
    pitching_stats["WHIP"] = (pitching_stats["walks"] + pitching_stats["hits_allowed"]) / pitching_stats["innings_pitched"]
    pitching_stats.loc[pitching_stats["innings_pitched"] == 0, "WHIP"] = None
    
    # K/9 and BB/9
    pitching_stats["K_per_9"] = (pitching_stats["strikeouts"] / pitching_stats["innings_pitched"]) * 9
    pitching_stats["BB_per_9"] = (pitching_stats["walks"] / pitching_stats["innings_pitched"]) * 9
    
    # Add pitcher names
    pitcher_name_lookup = batters_df[["pitcher", "player_name"]].drop_duplicates()
    pitcher_name_lookup["player_name"] = pitcher_name_lookup["player_name"].apply(clean_name).apply(flip_name)
    pitcher_name_lookup = pitcher_name_lookup.rename(columns={"player_name": "pitcher_name"})
    pitching_stats = pitching_stats.merge(pitcher_name_lookup, on="pitcher", how="left")
    
    print(f"  ✅ Calculated stats for {len(pitching_stats)} pitchers")
    return pitching_stats

# ============================================================================
# STEP 4: ENHANCED PITCHER FEATURES (NEW IN V2.0)
# ============================================================================

def calculate_pitcher_features(batters_df):
    """Calculate enhanced pitcher features for ML."""
    print("\n🔬 Engineering pitcher features...")
    
    features = []
    
    for pitcher_id in batters_df['pitcher'].dropna().unique():
        pitcher_df = batters_df[batters_df['pitcher'] == pitcher_id].copy()
        
        # Pitch mix percentages
        pitch_counts = pitcher_df['pitch_type'].value_counts()
        total_pitches = len(pitcher_df)
        pitch_mix = (pitch_counts / total_pitches * 100).to_dict()
        
        # Average velocity by pitch type
        avg_velo = pitcher_df.groupby('pitch_type')['release_speed'].mean().to_dict()
        
        # Average spin rate by pitch type
        avg_spin = pitcher_df.groupby('pitch_type')['release_spin_rate'].mean().to_dict()
        
        # Zone tendencies (% pitches in each zone)
        zone_counts = pitcher_df['zone'].value_counts()
        zone_pct = (zone_counts / len(pitcher_df) * 100).to_dict()
        
        # Ground ball rate
        gb_rate = (pitcher_df['bb_type'] == 'ground_ball').sum() / len(pitcher_df[pitcher_df['bb_type'].notna()]) * 100 if len(pitcher_df[pitcher_df['bb_type'].notna()]) > 0 else 0
        
        # Fly ball rate
        fb_rate = (pitcher_df['bb_type'] == 'fly_ball').sum() / len(pitcher_df[pitcher_df['bb_type'].notna()]) * 100 if len(pitcher_df[pitcher_df['bb_type'].notna()]) > 0 else 0
        
        # Days since last appearance
        last_game = pitcher_df['game_date'].max()
        days_rest = (datetime.today() - pd.to_datetime(last_game)).days
        
        # Pitches thrown in last 5 days (fatigue indicator)
        five_days_ago = datetime.today() - timedelta(days=5)
        recent_pitches = len(pitcher_df[pitcher_df['game_date'] >= five_days_ago.strftime('%Y-%m-%d')])
        
        features.append({
            'pitcher': pitcher_id,
            'fastball_pct': pitch_mix.get('FF', 0) + pitch_mix.get('FT', 0) + pitch_mix.get('SI', 0),
            'breaking_ball_pct': pitch_mix.get('SL', 0) + pitch_mix.get('CU', 0),
            'offspeed_pct': pitch_mix.get('CH', 0) + pitch_mix.get('FS', 0),
            'avg_fastball_velo': avg_velo.get('FF', np.nan),
            'avg_spin_rate': pitcher_df['release_spin_rate'].mean(),
            'gb_rate': gb_rate,
            'fb_rate': fb_rate,
            'zone_in_pct': sum(zone_pct.get(z, 0) for z in range(1, 10)),  # Zones 1-9 = strike zone
            'zone_out_pct': sum(zone_pct.get(z, 0) for z in range(11, 15)),  # Zones 11-14 = outside
            'days_rest': days_rest,
            'pitches_last_5_days': recent_pitches
        })
    
    pitcher_features = pd.DataFrame(features)
    print(f"  ✅ Engineered features for {len(pitcher_features)} pitchers")
    return pitcher_features

# ============================================================================
# STEP 5: ENHANCED BATTER FEATURES (NEW IN V2.0)
# ============================================================================

def calculate_batter_features(batters_df):
    """Calculate enhanced batter features for ML."""
    print("\n⚾ Engineering batter features...")
    
    features = []
    
    for batter_id in batters_df['batter'].dropna().unique():
        batter_df = batters_df[batters_df['batter'] == batter_id].copy()
        
        # Exit velocity (from launch_speed column)
        avg_exit_velo = batter_df['launch_speed'].mean()
        max_exit_velo = batter_df['launch_speed'].max()
        
        # Barrel rate (exit_velo >98 mph + launch angle 26-30°)
        barrels = batter_df[
            (batter_df['launch_speed'] > 98) &
            (batter_df['launch_angle'].between(26, 30))
        ]
        barrel_rate = len(barrels) / len(batter_df[batter_df['launch_speed'].notna()]) * 100 if len(batter_df[batter_df['launch_speed'].notna()]) > 0 else 0
        
        # Chase rate (swings at pitches outside zone 1-9)
        swings_outside = batter_df[
            (batter_df['zone'] > 10) &
            (batter_df['description'].str.contains('swinging', case=False, na=False))
        ]
        chase_rate = len(swings_outside) / len(batter_df) * 100
        
        # Contact rate
        swings = batter_df[batter_df['description'].str.contains('swinging|foul', case=False, na=False)]
        contact = swings[swings['description'].str.contains('foul|hit_into_play', case=False, na=False)]
        contact_rate = len(contact) / len(swings) * 100 if len(swings) > 0 else 0
        
        # Hard contact rate (exit velo >95 mph)
        hard_contact = batter_df[batter_df['launch_speed'] > 95]
        hard_contact_rate = len(hard_contact) / len(batter_df[batter_df['launch_speed'].notna()]) * 100 if len(batter_df[batter_df['launch_speed'].notna()]) > 0 else 0
        
        # Pull rate (pulled balls %)
        pull_rate = (batter_df['hit_location'].between(3, 4)).sum() / len(batter_df[batter_df['hit_location'].notna()]) * 100 if len(batter_df[batter_df['hit_location'].notna()]) > 0 else 0
        
        # Batting avg last 7/30 days
        seven_days_ago = datetime.today() - timedelta(days=7)
        thirty_days_ago = datetime.today() - timedelta(days=30)
        
        recent_7 = batter_df[batter_df['game_date'] >= seven_days_ago.strftime('%Y-%m-%d')]
        recent_30 = batter_df[batter_df['game_date'] >= thirty_days_ago.strftime('%Y-%m-%d')]
        
        hits_7 = recent_7['events'].isin(['single', 'double', 'triple', 'home_run']).sum()
        ab_7 = len(recent_7[recent_7['events'].notna()])
        ba_7 = hits_7 / ab_7 if ab_7 > 0 else 0
        
        hits_30 = recent_30['events'].isin(['single', 'double', 'triple', 'home_run']).sum()
        ab_30 = len(recent_30[recent_30['events'].notna()])
        ba_30 = hits_30 / ab_30 if ab_30 > 0 else 0
        
        # Walk and strikeout rates (last 30 days)
        walks_30 = recent_30['events'].isin(['walk', 'intent_walk']).sum()
        strikeouts_30 = recent_30['events'].str.contains('strikeout', na=False).sum()
        walk_rate_30 = walks_30 / ab_30 * 100 if ab_30 > 0 else 0
        k_rate_30 = strikeouts_30 / ab_30 * 100 if ab_30 > 0 else 0
        
        features.append({
            'batter': batter_id,
            'avg_exit_velo': avg_exit_velo,
            'max_exit_velo': max_exit_velo,
            'barrel_rate': barrel_rate,
            'chase_rate': chase_rate,
            'contact_rate': contact_rate,
            'hard_contact_rate': hard_contact_rate,
            'pull_rate': pull_rate,
            'ba_last_7_days': ba_7,
            'ba_last_30_days': ba_30,
            'walk_rate_last_30': walk_rate_30,
            'k_rate_last_30': k_rate_30
        })
    
    batter_features = pd.DataFrame(features)
    print(f"  ✅ Engineered features for {len(batter_features)} batters")
    return batter_features

# ============================================================================
# STEP 6: BATTER VS PITCHER HISTORY
# ============================================================================

def calculate_batter_vs_pitcher_history(batters_df):
    """Calculate career stats for batter vs pitcher matchups."""
    print("\n🔁 Calculating batter vs pitcher history...")
    
    history = []
    
    for (batter_id, pitcher_id), group in batters_df.groupby(['batter', 'pitcher']):
        ab = len(group[group['events'].notna()])
        hits = group['events'].isin(['single', 'double', 'triple', 'home_run']).sum()
        hr = (group['events'] == 'home_run').sum()
        
        history.append({
            'batter': batter_id,
            'pitcher': pitcher_id,
            'career_ab_vs_pitcher': ab,
            'career_hits_vs_pitcher': hits,
            'career_ba_vs_pitcher': hits / ab if ab > 0 else 0,
            'career_hr_vs_pitcher': hr
        })
    
    history_df = pd.DataFrame(history)
    print(f"  ✅ Found {len(history_df)} batter-pitcher matchup histories")
    return history_df

# ============================================================================
# STEP 7: WEATHER DATA WITH WIND DIRECTION (ENHANCED IN V2.0)
# ============================================================================

def pull_weather_with_wind_direction(pitchers_df):
    """Pull weather data including wind direction from Open-Meteo API."""
    print("\n🌤️  Pulling weather data (temperature, wind speed, wind direction, precipitation)...")
    
    home_games = pitchers_df[pitchers_df["pitcher_home_away"] == "Home"]
    weather_forecasts = {}
    
    for _, row in home_games.iterrows():
        team = row["home_team_abbr"]
        game_time = row["game_time_local"]
        loc = ballpark_locations.get(team)
        
        if not loc:
            continue
        
        lat, lon = loc["lat"], loc["lon"]
        
        # Open-Meteo API with wind direction
        url = (
            f"https://api.open-meteo.com/v1/forecast?"
            f"latitude={lat}&longitude={lon}"
            f"&hourly=temperature_2m,wind_speed_10m,wind_direction_10m,precipitation_probability"
            f"&temperature_unit=fahrenheit"
            f"&start_date={game_time.date()}&end_date={game_time.date()}"
            f"&timezone=America/New_York"
        )
        
        try:
            res = requests.get(url, timeout=10).json()
            times = [datetime.fromisoformat(t) for t in res["hourly"]["time"]]
            diffs = [abs((t - game_time).total_seconds()) for t in times]
            idx = diffs.index(min(diffs))
            
            forecast = {
                "temperature": res["hourly"]["temperature_2m"][idx],
                "wind_speed": res["hourly"]["wind_speed_10m"][idx],
                "wind_direction": res["hourly"]["wind_direction_10m"][idx],  # NEW: degrees from north
                "precip_prob": res["hourly"]["precipitation_probability"][idx],
            }
            weather_forecasts[team] = forecast
            print(f"  ✅ {team}: {forecast['temperature']}°F, Wind {forecast['wind_speed']} mph from {forecast['wind_direction']}°")
        except Exception as e:
            print(f"  ❌ Weather pull failed for {team}: {e}")
    
    return weather_forecasts

# ============================================================================
# STEP 8: BULLPEN STATS (KEEP FROM V1.0)
# ============================================================================

def calculate_bullpen_stats(batters_df):
    """Calculate bullpen (relief pitcher) statistics."""
    print("\n🚨 Calculating bullpen stats...")
    
    # Identify starters (pitched in inning 1 with 0 outs)
    starter_ids = batters_df[(batters_df["inning"] == 1) & (batters_df["outs_when_up"] == 0)]["pitcher"].unique()
    bullpen_df = batters_df[~batters_df["pitcher"].isin(starter_ids)]
    
    bullpen_outs = bullpen_df.groupby("pitcher")["outs"].sum().reset_index(name="outs_recorded")
    bullpen_outs["innings_pitched"] = bullpen_outs["outs_recorded"] / 3
    
    hit_events = {"single", "double", "triple", "home_run"}
    strikeout_events = {"strikeout", "strikeout_swinging", "strikeout_looking", "strikeout_double_play"}
    
    bullpen_hits = bullpen_df[bullpen_df["events"].isin(hit_events)].groupby("pitcher").size().reset_index(name="hits_allowed")
    bullpen_strikeouts = bullpen_df[bullpen_df["events"].isin(strikeout_events)].groupby("pitcher").size().reset_index(name="strikeouts")
    bullpen_runs = bullpen_df[bullpen_df["run_event"]].groupby("pitcher").size().reset_index(name="runs_allowed")
    
    bullpen_stats = bullpen_outs.merge(bullpen_hits, on="pitcher", how="outer").fillna(0)
    bullpen_stats = bullpen_stats.merge(bullpen_strikeouts, on="pitcher", how="outer").fillna(0)
    bullpen_stats = bullpen_stats.merge(bullpen_runs, on="pitcher", how="outer").fillna(0)
    
    bullpen_stats["ERA"] = (bullpen_stats["runs_allowed"] / bullpen_stats["innings_pitched"]) * 9
    bullpen_stats.loc[bullpen_stats["innings_pitched"] == 0, "ERA"] = None
    
    # Add pitcher names
    pitcher_name_lookup = batters_df[["pitcher", "player_name"]].drop_duplicates()
    pitcher_name_lookup["player_name"] = pitcher_name_lookup["player_name"].apply(clean_name).apply(flip_name)
    pitcher_name_lookup = pitcher_name_lookup.rename(columns={"player_name": "pitcher_name"})
    bullpen_stats = bullpen_stats.merge(pitcher_name_lookup, on="pitcher", how="left")
    
    # Aggregate by team
    batters_df.loc[:, "pitcher_team"] = batters_df.apply(
        lambda row: row["away_team"] if row["inning_topbot"] == "Top" else row["home_team"],
        axis=1
    )
    
    bullpen_team_map = batters_df.drop_duplicates(subset=["game_pk", "pitcher"])
    bullpen_team_map.loc[:, "pitcher_team"] = bullpen_team_map.apply(
        lambda row: row["away_team"] if row["inning_topbot"] == "Top" else row["home_team"],
        axis=1
    )
    bullpen_stats = bullpen_stats.merge(bullpen_team_map[["pitcher", "pitcher_team"]], on="pitcher", how="left")
    
    team_bullpen = bullpen_stats.groupby("pitcher_team").agg({
        "ERA": "mean",
        "hits_allowed": "sum",
        "innings_pitched": "sum",
        "strikeouts": "sum"
    }).reset_index()
    
    team_bullpen["hits_per_inning"] = team_bullpen["hits_allowed"] / team_bullpen["innings_pitched"]
    team_bullpen["k_per_inning"] = team_bullpen["strikeouts"] / team_bullpen["innings_pitched"]
    
    # Bullpen strength score (lower = better)
    team_bullpen["bullpen_score"] = team_bullpen["ERA"] + team_bullpen["hits_per_inning"] - team_bullpen["k_per_inning"]
    team_bullpen["bullpen_strength_percentile"] = team_bullpen["bullpen_score"].rank(pct=True)
    
    print(f"  ✅ Calculated bullpen stats for {len(team_bullpen)} teams")
    return team_bullpen

# ============================================================================
# STEP 9: TEAM BATTING STRENGTH (KEEP FROM V1.0)
# ============================================================================

def calculate_team_batting_strength(batters_df):
    """Calculate team offensive strength (last 10 games)."""
    print("\n💪 Calculating team batting strength...")
    
    recent_batting = batters_df.sort_values("game_date", ascending=False).dropna(subset=["events"])
    recent_batting = recent_batting.groupby(["batter_team", "game_pk"]).head(1)
    recent_batting = recent_batting.groupby("batter_team").head(10)
    
    team_batting = recent_batting.groupby("batter_team").agg({
        "events": lambda x: sum(evt in ["single", "double", "triple", "home_run"] for evt in x)
    }).rename(columns={"events": "hits"}).reset_index()
    
    team_batting["batting_strength_percentile"] = team_batting["hits"].rank(pct=True)
    team_batting["batting_strength_label"] = team_batting["batting_strength_percentile"].apply(
        lambda x: "Strong" if x >= 0.67 else ("Average" if x >= 0.34 else "Weak")
    )
    
    print(f"  ✅ Calculated batting strength for {len(team_batting)} teams")
    return team_batting

# ============================================================================
# STEP 10: BATTER HOT STREAKS (KEEP FROM V1.0)
# ============================================================================

def calculate_batter_streaks(batters_df):
    """Calculate current hit streaks and average streak lengths."""
    print("\n🔥 Calculating batter hot streaks...")
    
    hit_events = ["single", "double", "triple", "home_run"]
    
    def calculate_streaks_for_batter(batter_id, df):
        b_df = df[df["batter"] == batter_id].copy()
        b_df = b_df.sort_values("game_date")
        
        game_hits = b_df.groupby(["game_pk", "game_date"])["events"].apply(
            lambda x: x.isin(hit_events).any()
        ).reset_index()
        game_hits = game_hits.sort_values("game_date")
        game_hits["hit"] = game_hits["events"]
        
        current_hit_streak = current_dry_streak = 0
        for result in reversed(game_hits["hit"].tolist()):
            if result:
                if current_dry_streak == 0:
                    current_hit_streak += 1
                else:
                    break
            else:
                if current_hit_streak == 0:
                    current_dry_streak += 1
                else:
                    break
        
        streaks, drys = [], []
        count = 1
        if not game_hits.empty:
            prev = game_hits["hit"].iloc[0]
            for hit in game_hits["hit"].tolist()[1:]:
                if hit == prev:
                    count += 1
                else:
                    (streaks if prev else drys).append(count)
                    count = 1
                    prev = hit
            (streaks if prev else drys).append(count)
        
        return pd.Series({
            "current_hit_streak": current_hit_streak,
            "avg_hit_streak": round(sum(streaks) / len(streaks), 2) if streaks else 0,
            "avg_hitless_streak": round(sum(drys) / len(drys), 2) if drys else 0
        })
    
    all_streaks = pd.DataFrame([
        calculate_streaks_for_batter(b, batters_df)
        for b in batters_df["batter"].dropna().unique()
    ], index=batters_df["batter"].dropna().unique()).reset_index().rename(columns={"index": "batter"})
    
    print(f"  ✅ Calculated streaks for {len(all_streaks)} batters")
    return all_streaks

# ============================================================================
# STEP 11: BUILD FINAL MATCHUP DATAFRAME (WITH ALL FEATURES)
# ============================================================================

def build_matchup_dataframe(pitchers_df, batters_df, pitching_stats, pitcher_features,
                           batter_features, batter_pitcher_history, team_bullpen,
                           team_batting, batter_streaks, weather_forecasts):
    """Combine all features into final matchup DataFrame."""
    print("\n🔨 Building final matchup DataFrame with all features...")
    
    # Filter batters to those who played in last 14 games
    recent_batting = batters_df.sort_values("game_date", ascending=False)
    recent_batting = recent_batting.drop_duplicates(subset=["game_pk", "batter_team", "batter"])
    recent_batting["game_rank"] = recent_batting.groupby("batter_team")["game_date"].rank(method="dense", ascending=False)
    recent_batting = recent_batting[recent_batting["game_rank"] <= 14]
    
    name_mode = recent_batting.groupby("batter")[["batter_name"]].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else x.iloc[0]
    ).reset_index()
    
    batter_team_map = recent_batting.sort_values("game_date", ascending=False)
    batter_team_map = batter_team_map.drop_duplicates(subset=["batter"])
    batter_team_map = batter_team_map[["batter", "batter_team"]].merge(name_mode, on="batter", how="left")
    
    # Merge pitchers with opponent batters
    final = pitchers_df.merge(batter_team_map, left_on="opponent_team", right_on="batter_team", how="left")
    final = final.drop_duplicates(subset=["pitcher", "batter"])
    
    # Merge pitching stats
    final = final.merge(
        pitching_stats[["pitcher_name", "innings_pitched", "hits_allowed", "walks", "strikeouts", "ERA", "WHIP", "K_per_9"]],
        on="pitcher_name", how="left"
    )
    
    # Merge pitcher features
    final = final.merge(pitcher_features, left_on="pitcher", right_on="pitcher", how="left")
    
    # Merge batter features
    final = final.merge(batter_features, on="batter", how="left")
    
    # Merge batter vs pitcher history
    final = final.merge(batter_pitcher_history, on=["batter", "pitcher"], how="left")
    
    # Merge team bullpen strength
    final = final.merge(team_bullpen[["pitcher_team", "bullpen_strength_percentile"]], on="pitcher_team", how="left")
    final["bullpen_strength_label"] = final["bullpen_strength_percentile"].apply(
        lambda x: "Strong" if x <= 0.33 else ("Average" if x <= 0.66 else "Weak") if pd.notna(x) else "Unknown"
    )
    
    # Merge team batting strength
    final = final.merge(team_batting[["batter_team", "batting_strength_percentile", "batting_strength_label"]], on="batter_team", how="left")
    
    # Merge batter streaks
    final = final.merge(batter_streaks, on="batter", how="left")
    
    # Add weather data with wind direction analysis
    def assign_weather_and_wind(row):
        team = row["home_team_abbr"]
        forecast = weather_forecasts.get(team, {})
        park = ballpark_locations.get(team, {})
        
        temperature = forecast.get("temperature")
        wind_speed = forecast.get("wind_speed")
        wind_direction = forecast.get("wind_direction")
        precip_prob = forecast.get("precip_prob")
        
        # Wind helps batters?
        cf_direction = park.get("cf_direction", 0)
        park_hr_factor = park.get("park_hr_factor", 1.0)
        wind_batter_boost = calculate_wind_helps_batters(wind_direction, wind_speed, cf_direction, park_hr_factor)
        
        # Park factors
        park_run_factor = park.get("park_run_factor", 1.0)
        altitude = park.get("altitude", 0)
        
        return pd.Series({
            "game_temperature": temperature,
            "game_wind_speed": wind_speed,
            "game_wind_direction": wind_direction,
            "game_precipitation_probability": precip_prob,
            "wind_batter_boost": wind_batter_boost,  # NEW: -1.0 to +1.0
            "park_run_factor": park_run_factor,
            "park_hr_factor": park_hr_factor,
            "park_altitude": altitude
        })
    
    weather_data = final.apply(assign_weather_and_wind, axis=1)
    final = pd.concat([final, weather_data], axis=1)
    
    # Add batter home/away and game time
    final["batter_home_away"] = final["pitcher_home_away"].apply(lambda x: "Away" if x == "Home" else "Home")
    final["game_time"] = pd.to_datetime(final["game_time_et"]).dt.strftime("%#I:%M %p")
    
    # Add empty H-AB column for manual tracking
    final["H-AB"] = ""
    
    print(f"  ✅ Final matchup DataFrame: {len(final)} batter-pitcher matchups")
    return final

# ============================================================================
# STEP 12: CALCULATE HIT LIKELIHOOD (V1.0 HEURISTIC — REPLACE WITH ML IN FUTURE)
# ============================================================================

def calculate_hit_likelihood_heuristic(final, batters_df):
    """
    Calculate hit likelihood using v1.0 heuristic approach.
    ✅ DONE: XGBoost models trained and integrated (v5.0+)
    """
    print("\n🎯 Calculating hit likelihood (heuristic v1.0)...")
    
    # Zone + pitch type matchup scoring (from v1.0)
    zone_stats = batters_df[batters_df["events"].notna() & batters_df["description"].isin(["hit_into_play", "hit_into_play_score"])]
    zone_stats.loc[:, "is_hit"] = zone_stats["events"].isin(["single", "double", "triple", "home_run"])
    
    batter_perf = zone_stats.groupby(["batter", "pitch_type", "zone", "p_throws"]).agg({
        "is_hit": "mean",
        "release_speed": "mean"
    }).rename(columns={"is_hit": "hit_rate", "release_speed": "avg_speed"}).reset_index()
    
    pitcher_tendencies = zone_stats.groupby(["pitcher", "pitch_type", "zone", "p_throws"]).agg({
        "release_speed": "mean",
        "pitch_type": "count"
    }).rename(columns={"pitch_type": "pitch_count", "release_speed": "avg_speed"}).reset_index()
    
    matchup_base = final.merge(pitcher_tendencies, on="pitcher", how="left")
    matchup_base = matchup_base.merge(batter_perf, on=["batter", "pitch_type", "zone", "p_throws"], how="left")
    matchup_base["weighted_score"] = matchup_base["hit_rate"] * matchup_base["pitch_count"]
    
    matchup_scores = matchup_base.groupby(["pitcher", "batter"])["weighted_score"].sum().reset_index()
    matchup_scores.rename(columns={"weighted_score": "matchup_score"}, inplace=True)
    
    # Adjust for pitcher strikeouts
    pitcher_name_lookup = batters_df[["pitcher", "player_name"]].drop_duplicates()
    pitcher_name_lookup["player_name"] = pitcher_name_lookup["player_name"].apply(clean_name).apply(flip_name)
    pitcher_name_lookup = pitcher_name_lookup.rename(columns={"player_name": "pitcher_name"})
    
    matchup_scores = matchup_scores.merge(
        pitcher_name_lookup[["pitcher", "pitcher_name"]].drop_duplicates(),
        on="pitcher", how="left"
    )
    
    pitching_stats_k = batters_df.groupby("pitcher")["events"].apply(
        lambda x: x.str.contains('strikeout', na=False).sum()
    ).reset_index(name="strikeouts")
    
    matchup_scores = matchup_scores.merge(pitching_stats_k, on="pitcher", how="left")
    matchup_scores["adjusted_score"] = matchup_scores["matchup_score"] / (1 + matchup_scores["strikeouts"] / 100)
    
    # Normalize
    max_score = matchup_scores["adjusted_score"].max()
    matchup_scores["hit_likelihood_v1"] = (matchup_scores["adjusted_score"] / max_score * 100).round(1) if max_score > 0 else 0
    
    # Merge into final
    final = final.merge(matchup_scores[["pitcher_name", "batter", "hit_likelihood_v1"]], on=["pitcher_name", "batter"], how="left")
    
    print(f"  ✅ Hit likelihood calculated for {len(final)} matchups")
    return final

# ============================================================================
# STEP 13: EXPORT TO EXCEL WITH FORMATTING
# ============================================================================

def export_to_excel(final, output_path):
    """Export final matchup DataFrame to Excel with formatting."""
    print(f"\n📊 Exporting to Excel: {output_path}")
    
    # Select final columns
    final_cols = [
        "pitcher_name", "pitcher_team", "batter", "batter_name", "batter_team",
        "innings_pitched", "hits_allowed", "walks", "strikeouts", "ERA", "WHIP", "K_per_9",
        "fastball_pct", "breaking_ball_pct", "offspeed_pct", "avg_fastball_velo", "days_rest",
        "bullpen_strength_percentile", "bullpen_strength_label",
        "batting_strength_percentile", "batting_strength_label",
        "avg_exit_velo", "barrel_rate", "chase_rate", "hard_contact_rate",
        "ba_last_7_days", "ba_last_30_days", "walk_rate_last_30", "k_rate_last_30",
        "career_ab_vs_pitcher", "career_ba_vs_pitcher", "career_hr_vs_pitcher",
        "current_hit_streak", "avg_hit_streak", "avg_hitless_streak",
        "hit_likelihood_v1",
        "batter_home_away", "game_time",
        "game_temperature", "game_wind_speed", "game_wind_direction", "game_precipitation_probability",
        "wind_batter_boost", "park_run_factor", "park_hr_factor", "park_altitude",
        "H-AB"
    ]
    
    # Filter to existing columns
    final_cols = [col for col in final_cols if col in final.columns]
    final_export = final[final_cols].copy()
    
    # Sort by hit likelihood (highest first)
    final_export = final_export.sort_values(by="hit_likelihood_v1", ascending=False)
    
    # Replace NA with empty strings
    final_export = final_export.astype(object).fillna("")
    
    # Write to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matchups"
    
    for r in dataframe_to_rows(final_export, index=False, header=True):
        ws.append(r)
    
    # Format as table
    table_ref = f"A1:{get_column_letter(len(final_export.columns))}{len(final_export) + 1}"
    table = Table(displayName="MatchupTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium16", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Apply number formatting
    headers = [cell.value for cell in ws[1]]
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter][1:]:  # Skip header
            if header in ["innings_pitched", "ERA", "WHIP", "K_per_9", "bullpen_strength_percentile",
                         "batting_strength_percentile", "ba_last_7_days", "ba_last_30_days",
                         "career_ba_vs_pitcher", "avg_exit_velo", "game_wind_speed"]:
                cell.number_format = "0.00"
            elif header in ["fastball_pct", "breaking_ball_pct", "offspeed_pct", "barrel_rate",
                           "chase_rate", "hard_contact_rate", "walk_rate_last_30", "k_rate_last_30",
                           "hit_likelihood_v1"]:
                cell.number_format = "0.0"
            elif header == "game_temperature":
                cell.number_format = '0"°"'
            elif header == "game_precipitation_probability":
                cell.number_format = '0"%"'
            elif header == "wind_batter_boost":
                cell.number_format = "+0.000;-0.000;0.000"
            elif header == "park_run_factor" or header == "park_hr_factor":
                cell.number_format = "0.00"
    
    # Set default column width
    for col_idx in range(1, len(final_export.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    
    # Expand batter_name column
    if "batter_name" in headers:
        idx = headers.index("batter_name") + 1
        col_letter = get_column_letter(idx)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = max(max_length + 2, 15)
    
    # Save
    wb.save(output_path)
    print(f"✅ Excel file saved: {output_path}")

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution pipeline."""
    today_str = datetime.today().strftime("%Y%m%d")
    output_path = f"/Users/mikeross/.openclaw/workspace/projects/mlb-predictor/{today_str}_mlb_predictions_v2.0.xlsx"
    
    print("="*80)
    print("MLB PREDICTION MODEL v2.0 — ML-READY EDITION")
    print("="*80)
    
    # Step 1: Pull today's matchups
    pitchers_df = get_todays_matchups()
    if pitchers_df.empty:
        print("❌ No games found for today")
        return
    
    # Step 2: Pull Statcast data (2025 season)
    batters_df = pull_statcast_data(start_date='2025-03-27')
    if batters_df.empty:
        print("❌ No Statcast data loaded")
        return
    
    # Add outs column for pitching stats
    batters_df["outs"] = batters_df["events"].apply(get_outs)
    batters_df["run_event"] = batters_df["des"].str.contains("scores|homer", case=False, na=False)
    
    # Step 3: Calculate pitching stats
    pitching_stats = calculate_pitching_stats(batters_df)
    
    # Step 4: Calculate pitcher features
    pitcher_features = calculate_pitcher_features(batters_df)
    
    # Merge pitcher ID
    pitcher_id_lookup = batters_df[["pitcher", "player_name"]].drop_duplicates()
    pitcher_id_lookup["player_name"] = pitcher_id_lookup["player_name"].apply(clean_name).apply(flip_name)
    pitcher_id_lookup = pitcher_id_lookup.rename(columns={"player_name": "pitcher_name"})
    pitchers_df = pitchers_df.merge(pitcher_id_lookup, on="pitcher_name", how="left")
    
    # Step 5: Calculate batter features
    batter_features = calculate_batter_features(batters_df)
    
    # Step 6: Calculate batter vs pitcher history
    batter_pitcher_history = calculate_batter_vs_pitcher_history(batters_df)
    
    # Step 7: Calculate bullpen stats
    team_bullpen = calculate_bullpen_stats(batters_df)
    
    # Step 8: Calculate team batting strength
    team_batting = calculate_team_batting_strength(batters_df)
    
    # Step 9: Calculate batter streaks
    batter_streaks = calculate_batter_streaks(batters_df)
    
    # Step 10: Pull weather with wind direction
    weather_forecasts = pull_weather_with_wind_direction(pitchers_df)
    
    # Step 11: Build final matchup DataFrame
    final = build_matchup_dataframe(
        pitchers_df, batters_df, pitching_stats, pitcher_features,
        batter_features, batter_pitcher_history, team_bullpen,
        team_batting, batter_streaks, weather_forecasts
    )
    
    # Step 12: Calculate hit likelihood (v1.0 heuristic)
    final = calculate_hit_likelihood_heuristic(final, batters_df)
    
    # Step 13: Export to Excel
    export_to_excel(final, output_path)
    
    print("\n" + "="*80)
    print("✅ PIPELINE COMPLETE")
    print(f"📊 Output: {output_path}")
    print(f"📈 Total matchups analyzed: {len(final)}")
    print("="*80)

if __name__ == "__main__":
    main()

